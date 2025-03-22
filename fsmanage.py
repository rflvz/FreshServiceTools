import os
import pandas as pd
import requests
import argparse
from colorama import Fore, Style, init
import openpyxl
from openpyxl.styles import Alignment, Font, PatternFill

# Inicializar colorama
init(autoreset=True)

# Configuración de la API (Clave definida directamente en el código)
api_key = 'your_api_key'  # Reemplaza 'your_default_api_key' con tu clave real
subdomain = 'subdomain'  # Subdominio de Freshservice
base_url = f'https://{subdomain}.freshservice.com/api/v2/assets/'


# Función para obtener los componentes de un activo
def get_asset_components(asset_id):
    url = f'https://{subdomain}.freshservice.com/api/v2/assets/{asset_id}/components/'  # Endpoint para componentes
    try:
        response = requests.get(url, auth=(api_key, ''))
        response.raise_for_status()  # Lanza una excepción si el código de estado no es 200
        data = response.json()
        if 'components' in data:  # Verificar si la clave 'components' está presente
            return data['components']
        else:
            print(f"{Fore.YELLOW}Advertencia: No se encontraron componentes para el ID {asset_id}")
            return []
    except requests.exceptions.RequestException as e:
        print(f"{Fore.RED}Error al obtener componentes para el ID {asset_id}: {e}")
        return []

# Función para obtener los departamentos desde la API
def get_departments():
    url = f'https://{subdomain}.freshservice.com/api/v2/departments/'  # URL correcta para departamentos
    try:
        response = requests.get(url, auth=(api_key, ''))
        response.raise_for_status()  # Lanza una excepción si el código de estado no es 200
        data = response.json()
        if 'departments' in data:  # Verificar si la clave 'departments' está presente
            return {dept['id']: dept['name'] for dept in data['departments']}
        else:
            print(f"{Fore.YELLOW}Advertencia: No se encontraron departamentos.")
            return {}
    except requests.exceptions.RequestException as e:
        print(f"{Fore.RED}Error al obtener departamentos: {e}")
        return {}

# Función para obtener los datos de un activo
def get_asset_data(asset_id):
    url = f'https://{subdomain}.freshservice.com/api/v2/assets/{asset_id}'  # Endpoint para obtener datos del activo
    try:
        response = requests.get(url, auth=(api_key, ''))
        response.raise_for_status()  # Lanza una excepción si el código de estado no es 200
        data = response.json()
        if 'asset' in data:  # Verificar si la clave 'asset' está presente
            return data['asset']
        else:
            print(f"{Fore.YELLOW}Advertencia: No se encontraron datos para el ID {asset_id}")
            return None
    except requests.exceptions.RequestException as e:
        print(f"{Fore.RED}Error al obtener datos para el ID {asset_id}: {e}")
        return None

# Función para obtener los tipos de activos desde la API
def get_asset_types():
    url = f'https://{subdomain}.freshservice.com/api/v2/asset_types/'  # URL para obtener los tipos de activos
    try:
        response = requests.get(url, auth=(api_key, ''))
        response.raise_for_status()  # Lanza una excepción si el código de estado no es 200
        data = response.json()
        if 'asset_types' in data:  # Verificar si la clave 'asset_types' está presente
            return {asset_type['id']: asset_type['name'] for asset_type in data['asset_types']}
        else:
            print(f"{Fore.YELLOW}Advertencia: No se encontraron tipos de activos.")
            return {}
    except requests.exceptions.RequestException as e:
        print(f"{Fore.RED}Error al obtener tipos de activos: {e}")
        return {}

# Función para obtener el nombre del tipo de activo desde la API
def get_asset_type_name(asset_type_id):
    url = f'https://{subdomain}.freshservice.com/api/v2/asset_types/{asset_type_id}'  # URL para obtener el tipo de activo
    try:
        response = requests.get(url, auth=(api_key, ''))
        response.raise_for_status()  # Lanza una excepción si el código de estado no es 200
        data = response.json()
        if 'asset_type' in data:  # Verificar si la clave 'asset_type' está presente
            return data['asset_type'].get('name', 'Unknown')  # Devolver el nombre del tipo de activo
        else:
            return 'Unknown'
    except requests.exceptions.RequestException as e:
        print(f"{Fore.RED}Error al obtener el tipo de activo para ID {asset_type_id}: {e}")
        return 'Unknown'

# Función para obtener el nombre de la ubicación desde la API
def get_location_name(location_id):
    url = f'https://{subdomain}.freshservice.com/api/v2/locations/{location_id}'  # URL para obtener la ubicación
    try:
        response = requests.get(url, auth=(api_key, ''))
        response.raise_for_status()  # Lanza una excepción si el código de estado no es 200
        data = response.json()
        if 'location' in data:  # Verificar si la clave 'location' está presente
            return data['location'].get('name', 'Unknown')  # Devolver el nombre de la ubicación
        else:
            return 'Unknown'
    except requests.exceptions.RequestException as e:
        print(f"{Fore.RED}Error al obtener la ubicación para ID {location_id}: {e}")
        return 'Unknown'

# Función para obtener la información del usuario desde la API
def get_user_info(user_id):
    """
    Obtiene la información del usuario desde la API.

    :param user_id: ID del usuario.
    :return: Diccionario con el nombre, apellido y correo electrónico del usuario.
    """
    url = f'https://{subdomain}.freshservice.com/api/v2/requesters/{user_id}'  # URL para obtener la información del usuario
    try:
        response = requests.get(url, auth=(api_key, ''))
        response.raise_for_status()  # Lanza una excepción si el código de estado no es 200
        data = response.json()
        if 'requester' in data:  # Verificar si la clave 'requester' está presente
            user = data['requester']
            return {
                "first_name": user.get('first_name', 'Unknown'),
                "last_name": user.get('last_name', 'Unknown'),
                "primary_email": user.get('primary_email', 'Unknown')
            }
        else:
            return {"first_name": "Unknown", "last_name": "Unknown", "primary_email": "Unknown"}
    except requests.exceptions.RequestException as e:
        print(f"{Fore.RED}Error al obtener la información del usuario para ID {user_id}: {e}")
        return {"first_name": "Unknown", "last_name": "Unknown", "primary_email": "Unknown"}

# Función para obtener el sistema operativo de la máquina desde la API
def get_system_os(asset_id):
    """
    Obtiene el sistema operativo de la máquina desde la API.

    :param asset_id: ID del activo.
    :return: Nombre del sistema operativo o 'Unknown' si no se encuentra.
    """
    # Construir correctamente la URL para evitar problemas con caracteres especiales
    url = f'https://{subdomain}.freshservice.com/api/v2/assets?include=type_fields&filter="name:\'ASSET-{asset_id}\'"'
    try:
        response = requests.get(url, auth=(api_key, ''))
        response.raise_for_status()  # Lanza una excepción si el código de estado no es 200
        data = response.json()
        if 'assets' in data and len(data['assets']) > 0:  # Verificar si hay activos en la respuesta
            asset = data['assets'][0]
            # Asegurarse de que la clave 'type_fields' esté presente y obtener el sistema operativo
            type_fields = asset.get('type_fields', {})
            return type_fields.get('os_23001176139', 'Unknown')  # Extraer el sistema operativo
        else:
            print(f"{Fore.YELLOW}Advertencia: No se encontró el sistema operativo para el activo ID {asset_id}.")
            return 'Unknown'
    except requests.exceptions.RequestException as e:
        print(f"{Fore.RED}Error al obtener el sistema operativo para el activo ID {asset_id}: {e}")
        return 'Unknown'

# Función para obtener la dirección IP de la máquina desde la API
def get_machine_ip(asset_id):
    """
    Obtiene la dirección IP de la máquina desde la API.

    :param asset_id: ID del activo.
    :return: Dirección IP o 'Unknown' si no se encuentra.
    """
    url = f'https://{subdomain}.freshservice.com/api/v2/assets?include=type_fields&filter="name:\'ASSET-{asset_id}\'"'
    try:
        response = requests.get(url, auth=(api_key, ''))
        response.raise_for_status()  # Lanza una excepción si el código de estado no es 200
        data = response.json()
        if 'assets' in data and len(data['assets']) > 0:  # Verificar si hay activos en la respuesta
            asset = data['assets'][0]
            # Asegurarse de que la clave 'type_fields' esté presente y obtener la dirección IP
            type_fields = asset.get('type_fields', {})
            return type_fields.get('computer_ip_address_23001176139', 'Unknown')  # Extraer la dirección IP
        else:
            print(f"{Fore.YELLOW}Advertencia: No se encontró la dirección IP para el activo ID {asset_id}.")
            return 'Unknown'
    except requests.exceptions.RequestException as e:
        print(f"{Fore.RED}Error al obtener la dirección IP para el activo ID {asset_id}: {e}")
        return 'Unknown'

# Función para procesar los IDs de activos
def process_asset_ids(ids_input, exclude_input=None):
    ids = []
    if os.path.isfile(ids_input):  # Si es un archivo
        with open(ids_input, 'r') as file:
            ids = file.read().strip().split(',')
    else:  # Si es una lista de IDs separada por comas
        ids = ids_input.split(',')

    expanded_ids = []
    for id_part in ids:
        id_part = id_part.strip()
        if '-' in id_part:  # Detectar rangos como 143-145
            try:
                start, end = map(int, id_part.split('-'))
                expanded_ids.extend(range(start, end + 1))  # Expandir el rango
            except ValueError:
                print(f"{Fore.RED}Error: Rango inválido '{id_part}'. Ignorando.")
        elif id_part.isdigit():  # Si es un ID individual
            expanded_ids.append(int(id_part))
        else:
            print(f"{Fore.RED}Error: ID inválido '{id_part}'. Ignorando.")

    # Procesar exclusiones si se especifican
    if exclude_input:
        exclude_ids = process_asset_ids(exclude_input)  # Reutilizar la misma lógica para exclusiones
        expanded_ids = [id_ for id_ in expanded_ids if id_ not in exclude_ids]

    return expanded_ids

# Función común para obtener los datos iniciales de un activo
def fetch_asset_data(asset_id, include_asset_data, include_departments, include_asset_type, include_location, include_user):
    """
    Obtiene los datos iniciales de un activo desde la API y los procesa según las opciones habilitadas.

    :param asset_id: ID del activo.
    :param include_asset_data: Indica si se deben obtener todos los datos del activo.
    :param include_departments: Indica si se debe obtener el departamento asociado.
    :param include_asset_type: Indica si se debe obtener el tipo de activo asociado.
    :param include_location: Indica si se debe obtener la ubicación asociada.
    :param include_user: Indica si se debe obtener la información del usuario asociada.
    :return: Diccionario con los datos procesados del activo.
    """
    asset_data = None
    department_name = None
    asset_type_name = None
    location_name = None
    user_info = None

    # Obtener los datos del activo si se especifica -a o si son necesarios para -d, -t, -l o -u
    if include_asset_data or include_departments or include_asset_type or include_location or include_user:
        asset_data = get_asset_data(asset_id)
        if not asset_data:
            return None  # Si no se pueden obtener los datos del activo, devolver None

    # Obtener el nombre del departamento si se especifica -d
    if include_departments:
        department_id = asset_data.get('department_id') if asset_data else None
        if department_id:
            department_mapping = get_departments()
            department_name = department_mapping.get(department_id, 'Unknown')

    # Obtener el tipo de activo si se especifica -t
    if include_asset_type:
        asset_type_id = asset_data.get('asset_type_id') if asset_data else None
        if asset_type_id:
            asset_type_name = get_asset_type_name(asset_type_id)

    # Obtener la localización si se especifica -l
    if include_location:
        location_id = asset_data.get('location_id') if asset_data else None
        if location_id:
            location_name = get_location_name(location_id)

    # Obtener la información del usuario si se especifica -u
    if include_user:
        user_id = asset_data.get('user_id') if asset_data else None
        if user_id:
            user_info = get_user_info(user_id)

    return {
        "asset_data": asset_data,
        "department_name": department_name,
        "asset_type_name": asset_type_name,
        "location_name": location_name,
        "user_info": user_info
    }

def format_excel_file(file_path):
    """
    Aplica formato al archivo Excel generado:
    - Ajusta el ancho de las columnas al tamaño del contenido más largo.
    - Aplica un fondo azul oscuro con fuente blanca a la primera fila.
    - Aplica un fondo azul claro al resto de las filas.

    :param file_path: Ruta del archivo Excel a formatear.
    """
    wb = openpyxl.load_workbook(file_path)
    ws = wb.active

    # Estilos
    header_fill = PatternFill(start_color="003366", end_color="003366", fill_type="solid")  # Azul oscuro
    header_font = Font(color="FFFFFF", bold=True)  # Fuente blanca y negrita
    row_fill = PatternFill(start_color="D9E1F2", end_color="D9E1F2", fill_type="solid")  # Azul claro

    # Ajustar el ancho de las columnas
    for col in ws.columns:
        max_length = 0
        col_letter = col[0].column_letter  # Obtener la letra de la columna
        for cell in col:
            try:
                if cell.value:
                    max_length = max(max_length, len(str(cell.value)))
            except:
                pass
        adjusted_width = max_length + 2  # Añadir unos caracteres extra
        ws.column_dimensions[col_letter].width = adjusted_width

    # Aplicar estilos a la primera fila (encabezados)
    for cell in ws[1]:
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = Alignment(horizontal="center", vertical="center")

    # Aplicar estilos al resto de las filas
    for row in ws.iter_rows(min_row=2):
        for cell in row:
            cell.fill = row_fill

    # Guardar el archivo con los cambios
    wb.save(file_path)

# Función principal
def main(ids_input, exclude_input, components, output_file, verbose, include_departments, include_asset_data, include_asset_type, include_location, include_user, include_system_os, include_machine_ip):
    # Si se especifica -a, habilitar automáticamente -d, -t, -l, -u, -s y -n
    if include_asset_data:
        include_departments = True
        include_asset_type = True
        include_location = True
        include_user = True
        include_system_os = True
        include_machine_ip = True

    # Diccionario de traducción de abreviaturas a nombres reales de componentes
    component_translation = {
        "cpu": "Processor",
        "ram": "Memory",
        "hdd": "Logical Drive",
        "nic": "Network Adapter"
    }

    """
    Diccionario de componentes:
    Este diccionario traduce abreviaturas de componentes a sus nombres completos utilizados en la API.

    component_translation = {
        "cpu": "Processor",       # Procesador
        "ram": "Memory",          # Memoria RAM
        "hdd": "Logical Drive",   # Disco lógico
        "nic": "Network Adapter"  # Adaptador de red
    }

    Ejemplo:
    Si se especifica "-c cpu ram", se buscarán componentes de tipo "Processor" y "Memory".
    """

    # Validar la opción -c
    if components:
        components = [component_translation.get(c.lower(), c) for c in components]

    # Procesar los IDs de activos
    asset_ids = process_asset_ids(ids_input, exclude_input)

    if not asset_ids:
        print(f"{Fore.RED}Error: No se encontraron IDs válidos en la entrada proporcionada.")
        return

    # Si se especifica -o y no tiene extensión, agregar ".xlsx" por defecto
    if output_file and not os.path.splitext(output_file)[1]:
        output_file += ".xlsx"

    all_data = []
    for asset_id in asset_ids:
        # Obtener el sistema operativo si se especifica -s
        system_os = None
        if include_system_os:
            system_os = get_system_os(asset_id)

        # Obtener la dirección IP si se especifica -n
        machine_ip = None
        if include_machine_ip:
            machine_ip = get_machine_ip(asset_id)

        # Llamar a la función común para obtener los datos iniciales del activo
        asset_info = fetch_asset_data(asset_id, include_asset_data, include_departments, include_asset_type, include_location, include_user)
        if not asset_info and not (include_system_os or include_machine_ip or components):
            continue  # Si no se pueden obtener los datos del activo y no se usan -s, -n o -c, omitirlo

        asset_data = asset_info["asset_data"] if asset_info else None
        department_name = asset_info["department_name"] if asset_info else None
        asset_type_name = asset_info["asset_type_name"] if asset_info else None
        location_name = asset_info["location_name"] if asset_info else None
        user_info = asset_info["user_info"] if asset_info else None

        # Obtener los componentes del activo si se especifica -c
        if components is not None:  # Solo buscar componentes si se especifica -c
            components_data = get_asset_components(asset_id)
            if components_data:  # Solo procesar si hay componentes
                for component in components_data:
                    component_type = component.get('component_type', 'Unknown')
                    component_details = component.get('component_data', [])

                    # Filtrar por tipos de componentes si se especifican
                    if not components or component_type in components:
                        for detail in component_details:
                            detail['component_type'] = component_type  # Agregar el tipo de componente
                            detail['asset_id'] = asset_id  # Agregar el ID del activo
                            detail['component_name'] = component.get('name', 'Unknown')  # Agregar el nombre del componente
                            detail['component_status'] = component.get('status', 'Unknown')  # Agregar el estado del componente
                            if include_departments:
                                detail['department_name'] = department_name  # Agregar el nombre del departamento
                            if include_asset_type:
                                detail['asset_type'] = asset_type_name  # Agregar el tipo de activo
                            if include_location:
                                detail['location_name'] = location_name  # Agregar el nombre de la ubicación
                            if include_user and user_info:
                                detail['user_first_name'] = user_info.get('first_name')
                                detail['user_last_name'] = user_info.get('last_name')
                                detail['user_email'] = user_info.get('primary_email')
                            if include_system_os:
                                detail['system_os'] = system_os
                            if include_machine_ip:
                                detail['machine_ip'] = machine_ip
                            all_data.append(detail)
        else:
            # Si no se especifica -c, solo agregar información básica del activo
            basic_data = {"asset_id": asset_id}
            if include_departments:
                basic_data["department_name"] = department_name
            if include_asset_type:
                basic_data["asset_type"] = asset_type_name
            if include_location:
                basic_data["location_name"] = location_name
            if include_user and user_info:
                basic_data["user_first_name"] = user_info.get('first_name')
                basic_data["user_last_name"] = user_info.get('last_name')
                basic_data["user_email"] = user_info.get('primary_email')
            if include_system_os:
                basic_data["system_os"] = system_os
            if include_machine_ip:
                basic_data["machine_ip"] = machine_ip
            all_data.append(basic_data)

    # Crear un DataFrame
    if all_data:
        df = pd.DataFrame(all_data)

        # Guardar en archivo si se especifica
        if output_file:
            df.to_excel(output_file, index=False)
            print(f"{Fore.GREEN}Datos guardados en {output_file}")

            # Aplicar formato si el archivo es .xlsx
            if output_file.endswith(".xlsx"):
                format_excel_file(output_file)

        # Mostrar en pantalla si verbose es True
        if verbose:
            print(f"{Fore.CYAN}Datos obtenidos:")
            print(df)
    else:
        print(f"{Fore.RED}No se obtuvieron datos.")

if __name__ == "__main__":
    # Configurar argumentos de línea de comandos
    parser = argparse.ArgumentParser(
        description=f"""{Fore.CYAN}Herramienta para obtener componentes de activos de Freshservice.

Diccionario de componentes:
Este diccionario traduce abreviaturas de componentes a sus nombres completos utilizados en la API.

component_translation = {{
    "cpu": "Processor",       # Procesador
    "ram": "Memory",          # Memoria RAM
    "hdd": "Logical Drive",   # Disco lógico
    "nic": "Network Adapter"  # Adaptador de red
}}

Ejemplo:
Si se especifica "-c cpu ram", se buscarán componentes de tipo "Processor" y "Memory".
""",
        epilog=f"{Fore.YELLOW}Ejemplo de uso: python fscomp.py -i 143-150 -e 145,147 -c cpu ram -o output.xlsx -v false -d -a -t -l -u -s -n"
    )
    parser.add_argument(
        '-i', '--ids',
        required=True,
        help=f"{Fore.GREEN}IDs de los activos separados por comas (por ejemplo: 143,197,302) o un archivo .txt con los IDs."
    )
    parser.add_argument(
        '-e', '--exclude',
        help=f"{Fore.GREEN}IDs de los activos a excluir, separados por comas o como un rango (por ejemplo: 145,147 o 145-147)."
    )
    parser.add_argument(
        '-c', '--components',
        nargs='*',  # Permitir cero o más valores
        help=f"{Fore.GREEN}Especifica los tipos de componentes a incluir (por ejemplo: processor memory). Si no se especifica, no se mostrarán componentes."
    )
    parser.add_argument(
        '-o', '--output',
        help=f"{Fore.GREEN}Nombre del archivo Excel donde se guardarán los datos (por ejemplo: output.xlsx)."
    )
    parser.add_argument(
        '-v', '--verbose',
        type=lambda x: x.lower() == 'true',
        default=True,
        help=f"{Fore.GREEN}Muestra los datos en pantalla (por defecto: true). Usa -v false para desactivar."
    )
    parser.add_argument(
        '-d', '--departments',
        action='store_true',
        help=f"{Fore.GREEN}Incluye el nombre del departamento asociado a cada activo."
    )
    parser.add_argument(
        '-a', '--asset-data',
        action='store_true',
        help=f"{Fore.GREEN}Obtiene los datos del activo desde la API (por ejemplo: department_id)."
    )
    parser.add_argument(
        '-t', '--asset-type',
        action='store_true',
        help=f"{Fore.GREEN}Incluye el tipo de activo asociado a cada activo."
    )
    parser.add_argument(
        '-l', '--location',
        action='store_true',
        help=f"{Fore.GREEN}Incluye el nombre de la ubicación asociada a cada activo."
    )
    parser.add_argument(
        '-u', '--user',
        action='store_true',
        help=f"{Fore.GREEN}Incluye la información del usuario asociada a cada activo (nombre, apellido y correo electrónico)."
    )
    parser.add_argument(
        '-s', '--system-os',
        action='store_true',
        help=f"{Fore.GREEN}Incluye el sistema operativo de la máquina asociada a cada activo."
    )
    parser.add_argument(
        '-n', '--machine-ip',
        action='store_true',
        help=f"{Fore.GREEN}Incluye la dirección IP de la máquina asociada a cada activo."
    )
    args = parser.parse_args()

    # Ejecutar la función principal con los argumentos especificados
    main(args.ids, args.exclude, args.components, args.output, args.verbose, args.departments, args.asset_data, args.asset_type, args.location, args.user, args.system_os, args.machine_ip)
